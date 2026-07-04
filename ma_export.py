"""
Export des résultats de valorisation M&A101 en Excel (multi-feuilles, mise en forme
façon modèle bancaire) et en PDF (résumé + football field).
"""
from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from fpdf import FPDF, XPos, YPos

PURPLE = "6B2FBF"
INK = "111111"
GREY = "5B5F6B"
FONT_NAME = "Arial"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_font():
    return Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)


def _header_fill():
    return PatternFill("solid", start_color=PURPLE, end_color=PURPLE)


def _title(ws, text, row=1, span=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=True, size=14, color=PURPLE)


def _write_table(ws, start_row, headers, rows, col_formats=None, col_widths=None):
    """Écrit un tableau avec en-tête coloré, bordures et formats de nombre. Retourne la ligne suivante libre."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, row_vals in enumerate(rows, start=1):
        for j, val in enumerate(row_vals, start=1):
            c = ws.cell(row=start_row + i, column=j, value=val)
            c.font = Font(name=FONT_NAME, size=10, color=INK)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j > 1 else "left")
            if col_formats and j in col_formats:
                c.number_format = col_formats[j]
    if col_widths:
        for j, w in col_widths.items():
            ws.column_dimensions[get_column_letter(j)].width = w
    return start_row + len(rows) + 2


def _kv_block(ws, start_row, title, pairs, col_widths=(26, 18)):
    """Bloc clé/valeur simple (libellé + valeur), avec un titre de section coloré."""
    ws.cell(row=start_row, column=1, value=title).font = Font(name=FONT_NAME, bold=True, size=12, color=PURPLE)
    r = start_row + 1
    for label, value in pairs:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=10, color=GREY)
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.font = Font(name=FONT_NAME, size=10, bold=True, color=INK)
        r += 1
    ws.column_dimensions["A"].width = col_widths[0]
    ws.column_dimensions["B"].width = col_widths[1]
    return r + 1


def build_ma_excel_report(target: dict, acquirer: dict, dcf_result: dict, sensitivity_df,
                           comps_stats: dict, trans_stats: dict, synergies: dict, synergy_mode: str,
                           valuation_ranges: dict, offer: dict, deal: dict, tranches: list, blended_rate: float,
                           eps_result: dict, leverage: dict, currency: str = "M") -> bytes:
    """Classeur Excel complet : Résumé, DCF, Comparables, Transactions, Synergies, Football Field, Financement."""
    wb = Workbook()

    # ---------------------------------------------------------------- Résumé ---
    ws = wb.active
    ws.title = "Résumé"
    _title(ws, f"M&A101 — {target.get('name') or 'Cible'}", span=4)
    ws.cell(row=2, column=1, value="Valorisation et structuration de l'opération").font = Font(
        name=FONT_NAME, italic=True, size=10, color=GREY)

    r = _kv_block(ws, 4, "Société cible", [
        ("Nom", target.get("name") or "n.c."),
        (f"Chiffre d'affaires ({currency})", target.get("revenue")),
        (f"EBITDA ({currency})", target.get("ebitda")),
        (f"Dette nette ({currency})", target.get("net_debt")),
        ("Actions en circulation (M)", target.get("shares_outstanding")),
        ("Cours actuel", target.get("price")),
    ])
    r = _kv_block(ws, r, "Société acquéreuse", [
        ("Nom", acquirer.get("name") or "n.c."),
        (f"Résultat net ({currency})", acquirer.get("net_income")),
        (f"EBITDA ({currency})", acquirer.get("ebitda")),
        ("Actions en circulation (M)", acquirer.get("shares_outstanding")),
    ])
    r = _kv_block(ws, r, "Prix d'offre", [
        ("Cours actuel", target.get("price")),
        ("Prime totale", offer.get("total_premium")),
        ("Prix d'offre par action", offer.get("offer_price")),
    ])

    if valuation_ranges:
        ws.cell(row=r, column=1, value="Football field — synthèse des méthodes").font = Font(
            name=FONT_NAME, bold=True, size=12, color=PURPLE)
        rows = [[m, lo, mid, hi] for m, (lo, mid, hi) in valuation_ranges.items()]
        _write_table(ws, r + 1, ["Méthode", "Bas", "Médiane", "Haut"], rows,
                     col_formats={2: "#,##0.00", 3: "#,##0.00", 4: "#,##0.00"},
                     col_widths={1: 28, 2: 12, 3: 12, 4: 12})

    # ---------------------------------------------------------------- DCF ---
    ws = wb.create_sheet("DCF")
    _title(ws, "DCF — flux de trésorerie actualisés", span=6)
    r = 3
    headers = ["Année", "CA", "EBITDA", "EBIT", "NOPAT", "Capex", "ΔBFR", "FCF", "Facteur", "VA du FCF"]
    rows = [[row["year"], row["revenue"], row["ebitda"], row["ebit"], row["nopat"], row["capex"],
             row["delta_nwc"], row["fcf"], row["discount_factor"], row["pv_fcf"]] for row in dcf_result["rows"]]
    fmt = {c: "#,##0.0" for c in range(2, 9)}
    fmt[9] = "0.000"
    fmt[10] = "#,##0.0"
    r = _write_table(ws, r, headers, rows, col_formats=fmt,
                     col_widths={1: 8, 2: 11, 3: 11, 4: 11, 5: 11, 6: 10, 7: 10, 8: 11, 9: 10, 10: 12})

    r = _kv_block(ws, r, "Résultat", [
        ("Valeur terminale", dcf_result["terminal_value"]),
        ("VA de la valeur terminale", dcf_result["pv_terminal_value"]),
        ("Valeur d'entreprise", dcf_result["enterprise_value"]),
        ("Valeur des fonds propres", dcf_result["equity_value"]),
        ("Valeur par action", dcf_result["value_per_share"]),
    ])

    if sensitivity_df is not None:
        ws.cell(row=r, column=1, value="Sensibilité — valeur par action (WACC x croissance terminale)").font = Font(
            name=FONT_NAME, bold=True, size=12, color=PURPLE)
        r += 1
        ws.cell(row=r, column=1, value="WACC \\ g").font = _header_font()
        ws.cell(row=r, column=1).fill = _header_fill()
        for j, col in enumerate(sensitivity_df.columns, start=2):
            c = ws.cell(row=r, column=j, value=str(col))
            c.font = _header_font()
            c.fill = _header_fill()
            c.alignment = Alignment(horizontal="center")
        for i, idx in enumerate(sensitivity_df.index, start=1):
            ws.cell(row=r + i, column=1, value=str(idx)).font = Font(name=FONT_NAME, bold=True, size=10)
            for j, col in enumerate(sensitivity_df.columns, start=2):
                val = sensitivity_df.loc[idx, col]
                c = ws.cell(row=r + i, column=j, value=None if val != val else float(val))
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="center")
                c.border = BORDER

    # ---------------------------------------------------------------- Comparables ---
    if comps_stats:
        ws = wb.create_sheet("Comparables")
        _title(ws, "Comparables boursiers", span=4)
        rows = [["Bas (P25)", comps_stats["low"]["multiple"], comps_stats["low"]["per_share"]],
                ["Médiane", comps_stats["median"]["multiple"], comps_stats["median"]["per_share"]],
                ["Haut (P75)", comps_stats["high"]["multiple"], comps_stats["high"]["per_share"]]]
        _write_table(ws, 3, ["Scénario", "Multiple EV/EBITDA", "Valeur par action"], rows,
                     col_formats={2: "0.0\"x\"", 3: "#,##0.00"}, col_widths={1: 16, 2: 18, 3: 16})

    # ---------------------------------------------------------------- Transactions ---
    if trans_stats:
        ws = wb.create_sheet("Transactions")
        _title(ws, "Transactions précédentes", span=4)
        rows = [["Bas (P25)", trans_stats["low"]["multiple"], trans_stats["low"]["per_share"]],
                ["Médiane", trans_stats["median"]["multiple"], trans_stats["median"]["per_share"]],
                ["Haut (P75)", trans_stats["high"]["multiple"], trans_stats["high"]["per_share"]]]
        _write_table(ws, 3, ["Scénario", "Multiple EV/EBITDA", "Valeur par action"], rows,
                     col_formats={2: "0.0\"x\"", 3: "#,##0.00"}, col_widths={1: 16, 2: 18, 3: 16})

    # ---------------------------------------------------------------- Synergies ---
    if synergies:
        ws = wb.create_sheet("Synergies")
        _title(ws, f"Synergies ({synergy_mode})", span=6)
        if synergy_mode == "coûts":
            headers = ["Année", "% en régime", "Avant impôt", "Après impôt", "VA"]
            rows = [[row["year"], row["phase_in_pct"], row["pretax"], row["aftertax"], row["pv"]]
                    for row in synergies["rows"]]
            fmt = {2: "0%", 3: "#,##0.0", 4: "#,##0.0", 5: "#,##0.0"}
        else:
            headers = ["Année", "CA stand-alone", "CA avec synergies", "Amélioration EBIT", "NOPAT incrémental", "VA"]
            rows = [[row["year"], row["revenue_standalone"], row["revenue_with_synergies"],
                     row["ebit_improvement"], row["nopat_improvement"], row["pv"]] for row in synergies["rows"]]
            fmt = {c: "#,##0.0" for c in range(2, 7)}
        r = _write_table(ws, 3, headers, rows, col_formats=fmt, col_widths={c: 15 for c in range(1, 7)})
        _kv_block(ws, r, "VAN des synergies", [
            ("VA explicite", synergies["pv_explicit_sum"]),
            ("Valeur terminale", synergies["terminal_value"]),
            ("VA de la valeur terminale", synergies["pv_terminal_value"]),
            ("VAN totale", synergies["npv"]),
        ])

    # ---------------------------------------------------------------- Football Field ---
    if valuation_ranges:
        ws = wb.create_sheet("Football Field")
        _title(ws, "Football field — comparaison des méthodes", span=5)
        methods = list(valuation_ranges.keys())
        headers = ["Méthode", "Bas", "Médiane", "Haut", "Largeur (bas->haut)"]
        rows = []
        for m in methods:
            lo, mid, hi = valuation_ranges[m]
            rows.append([m, lo, mid, hi - lo])
        r = _write_table(ws, 3, headers, rows,
                         col_formats={2: "#,##0.00", 3: "#,##0.00", 4: "#,##0.00"},
                         col_widths={1: 26, 2: 12, 3: 12, 4: 16})

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Football field (valeur par action)"
        chart.y_axis.title = None
        chart.x_axis.title = "Valeur par action"
        data_start, data_end = 3, 3 + len(rows) - 1
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end + 1)
        low_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_end + 1)
        width_ref = Reference(ws, min_col=4, min_row=data_start, max_row=data_end + 1)
        chart.add_data(low_ref, titles_from_data=True)
        chart.add_data(width_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.noFill = True
        chart.series[1].graphicalProperties.solidFill = PURPLE
        chart.height, chart.width = 9, 18
        ws.add_chart(chart, f"A{r+1}")

    # ---------------------------------------------------------------- Financement & BPA ---
    ws = wb.create_sheet("Financement & BPA")
    _title(ws, "Structuration du financement", span=4)
    r = _kv_block(ws, 3, "Valeur du deal", [
        ("Valeur des fonds propres", deal.get("equity_value")),
        ("Valeur totale du deal", deal.get("total_deal_value")),
        ("Financé en cash", deal.get("cash_financed")),
        ("Financé par nouvelle dette", deal.get("debt_financed")),
    ])
    if tranches:
        ws.cell(row=r, column=1, value="Tranches de dette").font = Font(name=FONT_NAME, bold=True, size=12, color=PURPLE)
        r += 1
        rows = [[f"Tranche {i+1}", t["amount"], t["coupon"]] for i, t in enumerate(tranches)]
        r = _write_table(ws, r, ["Tranche", "Montant", "Coupon"], rows,
                         col_formats={2: "#,##0.0", 3: "0.00%"}, col_widths={1: 14, 2: 14, 3: 12})
    r = _kv_block(ws, r, "Coût de la dette", [("Taux moyen pondéré", blended_rate)])

    r = _kv_block(ws, r, "Impact sur le BPA de l'acquéreur", [
        ("BPA stand-alone", eps_result.get("standalone_eps")),
        ("BPA pro forma", eps_result.get("pro_forma_eps")),
        ("Accretion / dilution", eps_result.get("accretion_dilution_pct")),
    ])
    if leverage:
        r = _kv_block(ws, r, "Capacité d'endettement", [
            ("Levier avant deal", leverage.get("leverage_pre")),
            ("Dette nette après deal", leverage.get("net_debt_post")),
            ("EBITDA pro forma", leverage.get("pro_forma_ebitda")),
            ("Levier après deal", leverage.get("leverage_post")),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_ma_pdf_report(target: dict, offer: dict, valuation_ranges: dict, dcf_result: dict,
                         eps_result: dict, football_field_png: bytes = None) -> bytes:
    """Rapport PDF de synthèse M&A101 : valorisation, offre, impact BPA, football field."""
    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(107, 47, 191)
    pdf.cell(0, 12, f"Rapport de valorisation - {target.get('name') or 'Cible'}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, "Genere par M&A101 - Finance101", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)
    pdf.set_text_color(0, 0, 0)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(107, 47, 191)
    pdf.cell(0, 8, "Résultats clés", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 10.5)

    rows = [
        ("Cours actuel", f"{target.get('price', 0):,.2f}"),
        ("Prime totale", f"{offer.get('total_premium', 0)*100:.1f}%"),
        ("Prix d'offre par action", f"{offer.get('offer_price', 0):,.2f}"),
        ("Valeur par action (DCF)", f"{dcf_result.get('value_per_share', 0):,.2f}" if dcf_result else "n.a."),
        ("BPA pro forma acquéreur", f"{eps_result.get('pro_forma_eps', 0):,.2f}" if eps_result else "n.a."),
        ("Impact sur le BPA", f"{eps_result.get('accretion_dilution_pct', 0)*100:+.1f}%" if eps_result else "n.a."),
    ]
    for label, value in rows:
        pdf.cell(90, 7, label)
        pdf.cell(0, 7, value, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    if valuation_ranges:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(107, 47, 191)
        pdf.cell(0, 8, "Fourchettes de valorisation par méthode", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 10)
        for method, (lo, mid, hi) in valuation_ranges.items():
            pdf.cell(70, 7, method)
            pdf.cell(0, 7, f"{lo:,.2f}   -   {mid:,.2f}   -   {hi:,.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    if football_field_png:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(107, 47, 191)
        pdf.cell(0, 8, "Football field", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(0, 0, 0)
        _add_image_from_bytes(pdf, football_field_png, width=190)

    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.ln(6)
    pdf.multi_cell(0, 5, "Analyse générée automatiquement, à des fins indicatives. Ceci n'est pas un conseil "
                          "en investissement ni une opinion de valorisation formelle.")

    return bytes(pdf.output())


def _add_image_from_bytes(pdf: FPDF, img_bytes: bytes, width: float) -> None:
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(img_bytes)
        tmp_path = tmp.name
    try:
        pdf.image(tmp_path, x=10, w=width)
    finally:
        os.unlink(tmp_path)
